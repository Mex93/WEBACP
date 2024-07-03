from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_interval
from openpyxl.styles import (
    Alignment, Font
)
import os
import datetime


class CExcelLog:
    folder_name_log = "logs_users"
    folder_name_sql = "logs_sql"

    # СТИЛЬ ШРИФТА
    FONT_HEADER = Font(
        name='Calibri',
        size=21,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF000000'
    )
    FONT_DATA = Font(
        name='Arial',
        size=18,
        bold=False,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF000000'
    )
    alignment = Alignment(
        horizontal='center',
        vertical='center',
        text_rotation=0,
        wrap_text=False,
        shrink_to_fit=False,
        indent=0
    )

    def __init__(self):
        pass

    @classmethod
    def print_user_log(cls, log_text: str) -> bool:

        folder_its_ok = False
        if os.path.isdir(f"{cls.folder_name_log}"):
            folder_its_ok = True
        else:
            os.mkdir(cls.folder_name_log + "/")

            if os.path.isdir(f"{cls.folder_name_log}") is True:
                folder_its_ok = True
        ###
        if folder_its_ok is True:

            cdate = datetime.datetime.now()
            #
            day = cdate.day
            month = cdate.month
            year = cdate.year

            secs = cdate.second
            hours = cdate.hour
            mins = cdate.minute

            file_name = f"user_log_{day:02}.{month:02}.{year}.xlsx"

            is_file_exists = False
            file_name_with_patch = f"{cls.folder_name_log}/{file_name}"
            if os.path.exists(file_name_with_patch) is True:
                is_file_exists = True

            # try:
            if not is_file_exists:
                wb = Workbook()
                ws = wb.active
                ws.title = "Действия пользователя"
                ws.append(("Действие:", "Дата:"))

                # задаём ширину и фонт для шапки
                cell_range = ws['A1':'E1']
                for i in cell_range:
                    for i2 in i:
                        letter_adress = i2.coordinate

                        ws[letter_adress].font = cls.FONT_HEADER
                        ws[letter_adress].alignment = cls.alignment

                interval = get_column_interval("A", "E")
                for item in interval:
                    ws.column_dimensions[item].width = 60
            else:
                wb = load_workbook(file_name_with_patch)
                ws = wb.active

            ws.append((log_text, f"{hours:02}:{mins:02}:{secs:02} {day:02}:{month:02}:{year}"))

            # # Задаём фонт для столбцов с данными

            # TODO Понять как весь лист ебануть под один шрифт
            # Не получается ебануть строку, всё время она смещается вниз после стиля

            # cell_range = ws['A2':'D2500']
            # for i in cell_range:  # Можешь изменить min_row, чтобы начать с другой строки
            #     # Проходимся по каждой ячейке в строке
            #     for cell in i:
            #         # Применяем шрифт к ячейке
            #         cell.font = cls.FONT_HEADER
            #         cell.alignment = cls.alignment

            wb.save(f"{cls.folder_name_log}/{file_name}")
            wb.close()

            # except Exception as err:
            # print(f"Внимание! Ошибка лога: '{err}'.")
            # return False

            # if os.path.exists(f"{cls.folder_name}/{file_name}") is True:
            #     pass
            # else:
            return True


    @classmethod
    def print_sql_log(cls, user: str, action: str, log_text: str) -> bool:

        folder_its_ok = False
        if os.path.isdir(f"{cls.folder_name_sql}"):
            folder_its_ok = True
        else:
            os.mkdir(cls.folder_name_sql + "/")

            if os.path.isdir(f"{cls.folder_name_sql}") is True:
                folder_its_ok = True
        ###
        if folder_its_ok is True:

            cdate = datetime.datetime.now()
            #
            day = cdate.day
            month = cdate.month
            year = cdate.year

            secs = cdate.second
            hours = cdate.hour
            mins = cdate.minute

            file_name = f"sql_log_{day:02}.{month:02}.{year}.xlsx"

            is_file_exists = False
            file_name_with_patch = f"{cls.folder_name_sql}/{file_name}"
            if os.path.exists(file_name_with_patch) is True:
                is_file_exists = True

            # try:
            if not is_file_exists:
                wb = Workbook()
                ws = wb.active
                ws.title = "Сохранённые данные SQL"
                ws.append(("Пользователь:", "Действие:", "Данные:", "Дата:"))

                # задаём ширину и фонт для шапки
                cell_range = ws['A1':'E1']
                for i in cell_range:
                    for i2 in i:
                        letter_adress = i2.coordinate

                        ws[letter_adress].font = cls.FONT_HEADER
                        ws[letter_adress].alignment = cls.alignment

                interval = get_column_interval("A", "E")
                for item in interval:
                    ws.column_dimensions[item].width = 60
            else:
                wb = load_workbook(file_name_with_patch)
                ws = wb.active

            ws.append((user, action, log_text, f"{hours:02}:{mins:02}:{secs:02} {day:02}:{month:02}:{year}"))

            # # Задаём фонт для столбцов с данными

            # TODO Понять как весь лист ебануть под один шрифт
            # Не получается ебануть строку, всё время она смещается вниз после стиля

            # cell_range = ws['A2':'D2500']
            # for i in cell_range:  # Можешь изменить min_row, чтобы начать с другой строки
            #     # Проходимся по каждой ячейке в строке
            #     for cell in i:
            #         # Применяем шрифт к ячейке
            #         cell.font = cls.FONT_HEADER
            #         cell.alignment = cls.alignment

            wb.save(f"{cls.folder_name_sql}/{file_name}")
            wb.close()

            # except Exception as err:
            # print(f"Внимание! Ошибка лога: '{err}'.")
            # return False

            # if os.path.exists(f"{cls.folder_name}/{file_name}") is True:
            #     pass
            # else:
            return True
